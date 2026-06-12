import os
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side


ALLOWED_EXPORT_DIR = os.path.abspath("exports")


def _safe_path(file_path):
    abspath = os.path.abspath(file_path)
    if not abspath.startswith(os.path.abspath(".")):
        raise PermissionError("路径越权")
    return abspath


class ExcelHandler:

    @staticmethod
    def export(headers, rows, file_path):
        file_path = _safe_path(file_path)
        wb = openpyxl.Workbook()
        ws = wb.active
        font = Font(name="微软雅黑", size=11)
        align = Alignment(horizontal="center", vertical="center")
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(name="微软雅黑", size=11, bold=True)
            cell.alignment = align
            cell.border = border

        for row in rows:
            ws.append(row)
            for cell in ws[ws.max_row]:
                cell.font = font
                cell.alignment = align
                cell.border = border

        wb.save(file_path)

    @staticmethod
    def import_excel(file_path):
        file_path = _safe_path(file_path)
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            rows.append(list(row))
        return rows

    @staticmethod
    def export_to_response(headers, rows, filename="export.xlsx", sheet_name="Sheet1"):
        """Export data to an HttpResponse for download."""
        from django.http import HttpResponse
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        font = Font(name="微软雅黑", size=11)
        align = Alignment(horizontal="center", vertical="center")
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(name="微软雅黑", size=11, bold=True)
            cell.alignment = align
            cell.border = border

        for row in rows:
            ws.append(row)
            for cell in ws[ws.max_row]:
                cell.font = font
                cell.alignment = align
                cell.border = border

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
