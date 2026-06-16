from django.db import transaction
import openpyxl
from django.http import HttpResponse
from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from utils.response import APIResponse
from utils.permissions import HasPermission
from utils.excel import ExcelHandler
from .models import Permission, MenuPermissionRelation
from .serializers import PermissionSerializer


class PermissionViewSet(viewsets.ModelViewSet):
    queryset = Permission.objects.all()
    serializer_class = PermissionSerializer
    permission_key = "permission:list"
    permission_key_map = {
        "create": "permission:add",
        "update": "permission:edit",
        "destroy": "permission:delete",
        "batch": "permission:delete",
        "export": "permission:list",
        "status": "permission:edit",
        "sort": "permission:edit",
        "batch_sort": "permission:edit",
        "template": None,
        "import_permissions": "permission:add",
    }

    def get_permissions(self):
        return [IsAuthenticated(), HasPermission()]

    def get_queryset(self):
        qs = super().get_queryset()
        name = self.request.query_params.get("permission_name")
        status_val = self.request.query_params.get("status")
        if name:
            qs = qs.filter(permission_name__icontains=name)
        if status_val is not None and status_val != "":
            qs = qs.filter(status=int(status_val))
        return qs

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return APIResponse.success(data=serializer.data, message="新增成功")

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = self.get_serializer(queryset, many=True)
        return APIResponse.success(data=serializer.data)

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return APIResponse.success(data=serializer.data)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop("partial", False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return APIResponse.success(data=serializer.data, message="更新成功")

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.delete()
        return APIResponse.success(message="删除成功")

    @action(detail=True, methods=["get", "put"], url_path="menus")
    def menus(self, request, pk=None):
        instance = self.get_object()
        if request.method == "GET":
            menu_ids = MenuPermissionRelation.objects.filter(
                permission=instance
            ).values_list("menu_id", flat=True)
            return APIResponse.success(data=list(menu_ids))
        menu_ids = request.data.get("menuIds")
        if not isinstance(menu_ids, list):
            return APIResponse.error(message="menuIds 必须是数组")
        with transaction.atomic():
            MenuPermissionRelation.objects.filter(permission=instance).delete()
            if menu_ids:
                relations = [MenuPermissionRelation(permission=instance, menu_id=mid) for mid in menu_ids]
                MenuPermissionRelation.objects.bulk_create(relations)
        return APIResponse.success(message="绑定成功")

    @action(detail=False, methods=["delete"], url_path="batch")
    def batch(self, request):
        """批量删除 — DELETE /api/permission/batch"""
        ids = request.data.get("ids", [])
        if not ids:
            return APIResponse.error(message="ids 不能为空")
        Permission.objects.filter(id__in=ids).delete()
        return APIResponse.success(message="批量删除成功")

    @action(detail=False, methods=["get"], url_path="export")
    def export(self, request):
        """\u5bfc\u51fa\u6743\u9650"""
        from datetime import datetime
        perms = self.get_queryset().order_by("sort_order")
        headers = ["ID", "\u6743\u9650\u540d\u79f0", "\u6743\u9650\u6807\u8bc6", "\u6392\u5e8f", "\u72b6\u6001", "\u521b\u5efa\u65f6\u95f4"]
        rows = [
            [p.id, p.permission_name, p.permission_key, p.sort_order,
             "\u542f\u7528" if p.status else "\u7981\u7528", str(p.create_time)]
            for p in perms
        ]
        date_str = datetime.now().strftime("%Y%m%d")
        return ExcelHandler.export_to_response(headers, rows, f"permissions_{date_str}.xlsx", "\u6743\u9650\u5217\u8868")

    @action(detail=True, methods=["put"], url_path="status")
    def status(self, request, pk=None):
        """状态切换 — PUT /api/permission/<id>/status"""
        instance = self.get_object()
        status_val = request.data.get("status")
        if status_val not in (0, 1):
            return APIResponse.error(message="状态值无效")
        instance.status = status_val
        instance.save(update_fields=["status", "update_time"])
        return APIResponse.success(message="状态更新成功")

    @action(detail=True, methods=["put"], url_path="sort")
    def sort(self, request, pk=None):
        instance = self.get_object()
        instance.sort_order = request.data.get("sortOrder", 0)
        instance.save(update_fields=["sort_order", "update_time"])
        return APIResponse.success(message="排序更新成功")

    @action(detail=False, methods=["post"], url_path="batch-sort")
    def batch_sort(self, request):
        data = request.data
        if not isinstance(data, list):
            return APIResponse.error(message="请传入数组")
        instances = []
        for item in data:
            item_id = item.get("id")
            if not item_id:
                return APIResponse.error(message="每项需要 id 字段")
            instances.append(Permission(id=item_id, sort_order=item.get("sortOrder", 0)))
        Permission.objects.bulk_update(instances, ["sort_order"])
        return APIResponse.success(message="排序更新成功")

    @action(detail=False, methods=["get"], url_path="template")
    def template(self, request):
        """\u4e0b\u8f7d\u6743\u9650\u5bfc\u5165\u6a21\u677f"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "\u6743\u9650\u5bfc\u5165\u6a21\u677f"
        headers = ["\u6743\u9650\u540d\u79f0", "\u6743\u9650\u6807\u8bc6", "\u6392\u5e8f", "\u72b6\u6001(1\u542f\u7528/0\u7981\u7528)"]
        ws.append(headers)
        ws.append(["\u7528\u6237\u67e5\u8be2", "user:list", "0", "1"])
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="permission_template.xlsx"'
        wb.save(response)
        return response

    @action(detail=False, methods=["post"], url_path="import-permissions")
    def import_permissions(self, request):
        """\u5bfc\u5165\u6743\u9650"""
        file = request.FILES.get("file")
        if not file:
            return APIResponse.error(message="\u8bf7\u4e0a\u4f20\u6587\u4ef6")
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
        except Exception:
            return APIResponse.error(message="\u6587\u4ef6\u683c\u5f0f\u9519\u8bef")
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        if not rows:
            return APIResponse.error(message="\u6587\u4ef6\u5185\u5bb9\u4e3a\u7a7a")
        success, skipped, errors = 0, 0, []
        existing_keys = set(Permission.objects.values_list("permission_key", flat=True))
        for idx, row in enumerate(rows, start=2):
            perm_name = str(row[0]).strip() if row[0] else ""
            perm_key = str(row[1]).strip() if row[1] else ""
            sort_order = int(row[2]) if row[2] is not None else 0
            status = int(row[3]) if row[3] is not None else 1
            if not perm_key:
                errors.append(f"\u7b2c{idx}\u884c: \u6743\u9650\u6807\u8bc6\u4e3a\u7a7a")
                continue
            if perm_key in existing_keys:
                skipped += 1
                continue
            Permission.objects.create(
                permission_name=perm_name, permission_key=perm_key,
                sort_order=sort_order, status=status,
            )
            existing_keys.add(perm_key)
            success += 1
        return APIResponse.success(data={"success": success, "skipped": skipped, "errors": errors})
