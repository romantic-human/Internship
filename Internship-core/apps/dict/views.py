import openpyxl
from django.utils import timezone
from django.http import HttpResponse
from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from utils.response import APIResponse
from utils.permissions import HasPermission
from .models import DictType, DictData
from .serializers import DictTypeSerializer, DictDataSerializer


class DictTypeViewSet(viewsets.ModelViewSet):
    queryset = DictType.objects.all()
    serializer_class = DictTypeSerializer
    permission_key = "dict:type:list"
    permission_key_map = {
        "create": "dict:type:add",
        "update": "dict:type:edit",
        "destroy": "dict:type:delete",
        "batch": "dict:type:delete",
    }

    def get_permissions(self):
        if self.action in ("list", "retrieve", "by_type"):
            return [IsAuthenticated()]
        return [IsAuthenticated(), HasPermission()]

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        # 支持按名称筛选
        dict_name = request.query_params.get("dict_name")
        if dict_name:
            queryset = queryset.filter(dict_name__icontains=dict_name)
        dict_type = request.query_params.get("dict_type")
        if dict_type:
            queryset = queryset.filter(dict_type__icontains=dict_type)
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = self.get_serializer(queryset, many=True)
        return APIResponse.success(data=serializer.data)

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return APIResponse.success(data=serializer.data, message="新增成功")

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return APIResponse.success(data=serializer.data)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop("partial", False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return APIResponse.success(data=serializer.data, message="更新成功")

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        if DictData.objects.filter(dict_type=instance.dict_type).exists():
            return APIResponse.error(message="该字典类型下存在数据项，请先删除数据项")
        instance.delete()
        return APIResponse.success(message="删除成功")

    @action(detail=True, methods=["put"], url_path="status")
    def status(self, request, pk=None):
        instance = self.get_object()
        status_val = request.data.get("status")
        if status_val not in (0, 1):
            return APIResponse.error(message="状态值无效")
        instance.status = status_val
        instance.save()
        return APIResponse.success(message="状态更新成功")

    @action(detail=False, methods=["delete"], url_path="batch")
    def batch(self, request):
        ids = request.data.get("ids", [])
        if not ids:
            return APIResponse.error(message="ids 不能为空")
        # 检查是否有子数据
        types_with_data = DictType.objects.filter(
            id__in=ids,
            dict_type__in=DictData.objects.values_list("dict_type", flat=True).distinct()
        ).values_list("dict_name", flat=True)
        if types_with_data:
            return APIResponse.error(
                message=f"以下字典类型下存在数据项，无法删除：{', '.join(types_with_data)}"
            )
        DictType.objects.filter(id__in=ids).delete()
        return APIResponse.success(message="批量删除成功")

    @action(detail=False, methods=["get"])
    def export(self, request):
        import csv
        from django.http import HttpResponse
        queryset = DictType.objects.all().order_by("-create_time")
        response = HttpResponse(content_type="text/csv; charset=utf-8-sig")
        response["Content-Disposition"] = (
            f"attachment; filename=dict_types_{timezone.now().strftime('%Y%m%d')}.csv"
        )
        writer = csv.writer(response)
        writer.writerow(["字典名称", "字典类型编码", "状态", "备注", "创建时间"])
        for d in queryset:
            writer.writerow([
                d.dict_name, d.dict_type,
                "启用" if d.status == 1 else "禁用",
                d.remark, d.create_time.strftime("%Y-%m-%d %H:%M:%S"),
            ])
        return response

    @action(detail=False, methods=["get"], url_path="template")
    def template(self, request):
        from django.http import HttpResponse
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "DictType Template"
        headers = ["Dict Name", "Dict Code", "Status(1/0)", "Remark"]
        ws.append(headers)
        ws.append(["User Gender", "sys_user_gender", "1", "Gender dict"])
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="dict_type_template.xlsx"'
        wb.save(response)
        return response

    @action(detail=False, methods=["post"], url_path="import-types")
    def import_types(self, request):
        file = request.FILES.get("file")
        if not file:
            return APIResponse.error(message="Please upload a file")
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
        except Exception:
            return APIResponse.error(message="Invalid file format")
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        if not rows:
            return APIResponse.error(message="File is empty")
        success, skipped, errors = 0, 0, []
        existing_types = set(DictType.objects.values_list("dict_type", flat=True))
        for idx, row in enumerate(rows, start=2):
            dict_name = str(row[0]).strip() if row[0] else ""
            dict_type = str(row[1]).strip() if row[1] else ""
            status = int(row[2]) if row[2] is not None else 1
            remark = str(row[3]).strip() if row[3] else ""
            if not dict_type:
                errors.append(f"Row {idx}: dict code is empty")
                continue
            if dict_type in existing_types:
                skipped += 1
                continue
            DictType.objects.create(dict_name=dict_name, dict_type=dict_type, status=status, remark=remark)
            existing_types.add(dict_type)
            success += 1
        return APIResponse.success(data={"success": success, "skipped": skipped, "errors": errors})



class DictDataViewSet(viewsets.ModelViewSet):
    queryset = DictData.objects.select_related("dict_type").all()
    serializer_class = DictDataSerializer
    permission_key = "dict:data:list"
    permission_key_map = {
        "create": "dict:data:add",
        "update": "dict:data:edit",
        "destroy": "dict:data:delete",
        "batch": "dict:data:delete",
    }

    def get_permissions(self):
        if self.action in ("list", "retrieve", "by_type"):
            return [IsAuthenticated()]
        return [IsAuthenticated(), HasPermission()]

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        # 按字典类型筛选
        dict_type = request.query_params.get("dict_type")
        if dict_type:
            queryset = queryset.filter(dict_type=dict_type)
        # 按标签筛选
        dict_label = request.query_params.get("dict_label")
        if dict_label:
            queryset = queryset.filter(dict_label__icontains=dict_label)
        # 按状态筛选
        status = request.query_params.get("status")
        if status is not None and status != "":
            try:
                queryset = queryset.filter(status=int(status))
            except (ValueError, TypeError):
                pass
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = self.get_serializer(queryset, many=True)
        return APIResponse.success(data=serializer.data)

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return APIResponse.success(data=serializer.data, message="新增成功")

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return APIResponse.success(data=serializer.data)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop("partial", False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return APIResponse.success(data=serializer.data, message="更新成功")

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.delete()
        return APIResponse.success(message="删除成功")

    @action(detail=True, methods=["put"], url_path="status")
    def status(self, request, pk=None):
        instance = self.get_object()
        status_val = request.data.get("status")
        if status_val not in (0, 1):
            return APIResponse.error(message="状态值无效")
        instance.status = status_val
        instance.save()
        return APIResponse.success(message="状态更新成功")

    @action(detail=False, methods=["delete"], url_path="batch")
    def batch(self, request):
        ids = request.data.get("ids", [])
        if not ids:
            return APIResponse.error(message="ids 不能为空")
        DictData.objects.filter(id__in=ids).delete()
        return APIResponse.success(message="批量删除成功")

    @action(detail=False, methods=["get"], url_path="type/(?P<dict_type>[^/.]+)")
    def by_type(self, request, dict_type=None):
        """根据字典类型编码获取所有启用的数据项（不分页，用于下拉框）"""
        queryset = DictData.objects.filter(
            dict_type=dict_type, status=1
        ).select_related("dict_type").order_by("sort_order")
        serializer = self.get_serializer(queryset, many=True)
        return APIResponse.success(data=serializer.data)

    @action(detail=False, methods=["get"], url_path="data-template")
    def data_template(self, request):
        """\u4e0b\u8f7d\u5b57\u5178\u6570\u636e\u5bfc\u5165\u6a21\u677f"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "\u5b57\u5178\u6570\u636e\u6a21\u677f"
        headers = ["\u5b57\u5178\u7f16\u7801", "\u6570\u636e\u6807\u7b7e", "\u6570\u636e\u503c", "\u6392\u5e8f", "\u72b6\u6001(1\u542f\u7528/0\u7981\u7528)"]
        ws.append(headers)
        ws.append(["sys_user_gender", "\u7537", "1", "0", "1"])
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="dict_data_template.xlsx"'
        wb.save(response)
        return response

    @action(detail=False, methods=["post"], url_path="import-data")
    def import_data(self, request):
        """\u5bfc\u5165\u5b57\u5178\u6570\u636e"""
        file = request.FILES.get("file")
        if not file:
            return APIResponse.error(message="\u8bf7\u4e0a\u4f20\u6587\u4ef6")
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
        except Exception:
            return APIResponse.error(message="\u6587\u4ef6\u683c\u5f0f\u9519\u8bef")
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        if not rows:
            return APIResponse.error(message="\u6587\u4ef6\u5185\u5bb9\u4e3a\u7a7a")
        success, skipped, errors = 0, 0, []
        for idx, row in enumerate(rows, start=2):
            dict_type = str(row[0]).strip() if row[0] else ""
            dict_label = str(row[1]).strip() if row[1] else ""
            dict_value = str(row[2]).strip() if row[2] else ""
            sort_order = int(row[3]) if row[3] is not None else 0
            status = int(row[4]) if row[4] is not None else 1
            if not dict_type:
                errors.append(f"\u7b2c{idx}\u884c: \u5b57\u5178\u7f16\u7801\u4e3a\u7a7a")
                continue
            if DictData.objects.filter(dict_type=dict_type, dict_value=dict_value).exists():
                skipped += 1
                continue
            DictData.objects.create(
                dict_type=dict_type, dict_label=dict_label,
                dict_value=dict_value, sort_order=sort_order, status=status,
            )
            success += 1
        return APIResponse.success(data={"success": success, "skipped": skipped, "errors": errors})
