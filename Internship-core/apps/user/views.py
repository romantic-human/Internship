"""用户模块视图 — 参考《组织架构模块设计方案.md》第 5.2 节"""

import os
import uuid

from django.conf import settings
from django.db import IntegrityError, transaction
from django.utils import timezone

from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.throttling import AnonRateThrottle
from rest_framework_simplejwt.tokens import RefreshToken, TokenError
from django.http import HttpResponse

from utils import APIResponse, HasPermission
from .models import User, UserRoleRelation
from .serializers import (

    LoginSerializer,

    UserSerializer,

    UserListSerializer,

    ChangePasswordSerializer,
    RefreshTokenSerializer,

    UserCreateSerializer,

    UserProfileSerializer,

)

class UserViewSet(viewsets.ModelViewSet):

    """用户管理 CRUD"""

    queryset = User.objects.select_related("department").prefetch_related("userrolerelation_set__role").all()

    serializer_class = UserSerializer
    permission_key = "user:list"
    permission_key_map = {
        "create": "user:add",
        "update": "user:edit",
        "destroy": "user:delete",
        "batch": "user:delete",
        "status": "user:edit",
        "reset_password": "user:edit",
        "update_password": "user:edit",
        "avatar": "user:edit",
        "profile": "user:edit",
        "export": "user:export",
        "import": "user:import",
    }

    def get_permissions(self):
        if self.action in ("login", "register", "refresh_token"):
            return [AllowAny()]
        return [IsAuthenticated(), HasPermission()]

    def get_serializer_class(self):
        if self.action == "create":
            return UserCreateSerializer
        if self.action == "list":
            return UserListSerializer
        return UserSerializer

    def create(self, request, *args, **kwargs):
        """新增用户 — 返回标准格式"""
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return APIResponse.success(data=serializer.data, message="新增成功")

    def update(self, request, *args, **kwargs):
        """编辑用户 — 返回标准格式"""
        partial = kwargs.pop("partial", False)
        instance = self.get_object()
        role_ids = request.data.get("role_ids")
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        if role_ids is not None and isinstance(role_ids, list):
            with transaction.atomic():
                instance.userrolerelation_set.exclude(role_id__in=role_ids).delete()
                existing = set(instance.userrolerelation_set.values_list("role_id", flat=True))
                new_ids = [rid for rid in role_ids if rid not in existing]
                if new_ids:
                    UserRoleRelation.objects.bulk_create(
                        [UserRoleRelation(user=instance, role_id=rid) for rid in new_ids],
                        ignore_conflicts=True,
                    )
        return APIResponse.success(data=serializer.data, message="更新成功")

    def destroy(self, request, *args, **kwargs):
        """删除用户 — 返回标准格式"""
        instance = self.get_object()
        instance.delete()
        return APIResponse.success(message="删除成功")

    def list(self, request, *args, **kwargs):
        """用户列表（分页 + 搜索过滤）"""
        queryset = self.filter_queryset(self.get_queryset())
        username = request.query_params.get("username", "").strip()
        status_val = request.query_params.get("status")
        department_id = request.query_params.get("department_id")
        role_id = request.query_params.get("role_id")
        start_date = request.query_params.get("start_date") or request.query_params.get("startDate")
        end_date = request.query_params.get("end_date") or request.query_params.get("endDate")
        if username:
            queryset = queryset.filter(username__icontains=username)
        if status_val is not None and status_val != "":
            queryset = queryset.filter(status=int(status_val))
        if department_id:
            queryset = queryset.filter(department_id=int(department_id))
        if role_id:
            queryset = queryset.filter(userrolerelation__role_id=int(role_id)).distinct()
        if start_date:
            queryset = queryset.filter(create_time__gte=start_date)
        if end_date:
            queryset = queryset.filter(create_time__lte=end_date)
        queryset = queryset.order_by("-create_time")
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = self.get_serializer(queryset, many=True)
        return APIResponse.success(data={"records": serializer.data, "total": queryset.count()})

    @action(
        detail=False,
        methods=["post"],
        permission_classes=[AllowAny],
        throttle_classes=[AnonRateThrottle],
    )
    def login(self, request):

        """

        用户登录 — POST /api/user/login

        Request:  { "username": "admin", "password": "admin123" }

        Response: { "code": 200, "data": { "access_token", "refresh_token", "user" } }

        """

        serializer = LoginSerializer(data=request.data)

        if not serializer.is_valid():

            return APIResponse.error(

                message="参数错误：用户名和密码不能为空", code=2000, http_status=400

            )

        username = serializer.validated_data["username"]

        raw_password = serializer.validated_data["password"]

        # 查找用户

        user = User.objects.filter(username=username).first()

        if not user:

            return APIResponse.error(

                message="用户名或密码错误", code=3000, http_status=401

            )

        # 校验密码 (bcrypt)

        if not user.check_password(raw_password):

            return APIResponse.error(

                message="用户名或密码错误", code=3000, http_status=401

            )

        # 校验状态

        if user.status != 1:

            return APIResponse.error(

                message="该账号已被禁用，请联系管理员", code=2003, http_status=400

            )

        # 更新最后登录时间

        user.last_login = timezone.now()

        user.save(update_fields=["last_login"])

        # 生成 JWT

        refresh = RefreshToken.for_user(user)

        return APIResponse.success(

            message="登录成功",

            data={

                "access_token": str(refresh.access_token),

                "refresh_token": str(refresh),

                "user": {

                    "id": user.id,

                    "username": user.username,

                    "nickname": user.nickname,

                    "avatar": user.avatar,

                    "roles": user.role_list,

                    "permissions": user.permission_list,

                },

            },

        )

    @action(
        detail=False,
        methods=["post"],
        permission_classes=[AllowAny],
        throttle_classes=[AnonRateThrottle],
    )
    def register(self, request):

        """

        用户注册 — POST /api/user/register

        Request:  { "username": "admin", "password": "admin123", "nickname": "管理员" }

        Response: { "code": 200, "data": { id, username, nickname, ... } }

        """

        serializer = UserCreateSerializer(data=request.data)

        if not serializer.is_valid():

            return APIResponse.error(

                message="参数错误", code=2000, http_status=400,

            )

        try:

            user = serializer.save()

        except IntegrityError:

            return APIResponse.conflict(message="用户名已存在")

        # 注册后自动登录，返回 token

        refresh = RefreshToken.for_user(user)

        return APIResponse.success(

            message="注册成功",

            data={

                "access_token": str(refresh.access_token),

                "refresh_token": str(refresh),

                "user": {

                    "id": user.id,

                    "username": user.username,

                    "nickname": user.nickname,

                    "avatar": user.avatar,

                    "roles": user.role_list,

                    "permissions": user.permission_list,

                },

            },

        )

    @action(detail=False, methods=["post"], permission_classes=[AllowAny], url_path="refresh-token")

    def refresh_token(self, request):

        """

        刷新 Token — POST /api/user/refresh-token

        Request:  { "refresh": "eyJ..." }

        Response: { "code": 200, "data": { "access_token": "eyJ...", "refresh_token": "eyJ..." } }

        """

        serializer = RefreshTokenSerializer(data=request.data)

        if not serializer.is_valid():

            return APIResponse.error(

                message="refresh_token 不能为空", code=2000, http_status=400,

            )

        try:

            refresh = RefreshToken(serializer.validated_data["refresh"])

            data = {

                "access_token": str(refresh.access_token),

                "refresh_token": str(refresh),

            }

            return APIResponse.success(data=data, message="Token 刷新成功")

        except TokenError:

            return APIResponse.error(

                message="refresh_token 无效或已过期，请重新登录",

                code=3002, http_status=401,

            )

    @action(detail=True, methods=["put"])

    def status(self, request, pk=None):

        """修改状态 — PUT /api/user/:id/status"""

        user = self.get_object()

        status_val = request.data.get("status")

        if status_val not in (0, 1):

            return APIResponse.error(message="状态值无效", code=2000, http_status=400)

        user.status = status_val

        user.save(update_fields=["status", "update_time"])

        return APIResponse.success(message="状态更新成功")

    @action(detail=False, methods=["put"], url_path="reset-password")
    def reset_password(self, request):
        """重置密码 — PUT /api/user/reset-password"""
        user_id = request.data.get("userId")
        new_password = request.data.get("password", "123456")
        if not user_id:
            return APIResponse.error(message="userId 不能为空", code=2000, http_status=400)
        try:
            user = User.objects.get(id=user_id)
        except User.DoesNotExist:
            return APIResponse.error(message="用户不存在", code=2004, http_status=404)
        user.set_password(new_password)
        user.save(update_fields=["password", "update_time"])
        return APIResponse.success(message="密码已重置")

    @action(detail=False, methods=["delete"], url_path="batch")
    def batch(self, request):
        """批量删除 — DELETE /api/user/batch"""
        ids = request.data.get("ids", [])
        if not ids:
            return APIResponse.error(message="ids 不能为空")
        User.objects.filter(id__in=ids).delete()
        return APIResponse.success(message="批量删除成功")


    @action(detail=False, methods=["put"], url_path="update-password")

    def update_password(self, request):

        """

        修改自己的密码 — PUT /api/user/update-password

        Request:  { "old_password": "old123", "new_password": "new123" }

        Response: { "code": 200, "message": "密码修改成功" }

        """

        serializer = ChangePasswordSerializer(data=request.data)

        if not serializer.is_valid():

            return APIResponse.error(

                message="参数错误", code=2000, http_status=400,

            )

        user = request.user

        if not user or not user.is_authenticated:

            return APIResponse.error(

                message="未登录", code=3000, http_status=401,

            )

        # 校验旧密码

        if not user.check_password(serializer.validated_data["old_password"]):

            return APIResponse.error(

                message="旧密码不正确", code=2000, http_status=400,

            )

        # 设置新密码

        user.set_password(serializer.validated_data["new_password"])

        user.save(update_fields=["password", "update_time"])

        return APIResponse.success(message="密码修改成功")

    @action(detail=False, methods=["get"], url_path="check-unique")
    def check_unique(self, request):
        """检查用户名/邮箱/手机号唯一性 — GET /api/user/check-unique?field=username&value=xxx&exclude_id=1"""
        field = request.query_params.get("field", "")
        value = request.query_params.get("value", "")
        exclude_id = request.query_params.get("exclude_id")
        if field not in ("username", "email", "telephone") or not value:
            return APIResponse.error(message="参数错误")
        qs = User.objects.filter(**{field: value})
        if exclude_id:
            qs = qs.exclude(id=int(exclude_id))
        exists = qs.exists()
        field_labels = {"username": "用户名", "email": "邮箱", "telephone": "手机号"}
        if exists:
            return APIResponse.success(data={"unique": False}, message=f"{field_labels[field]}已存在")
        return APIResponse.success(data={"unique": True})

    @action(detail=False, methods=["get"], url_path="template")
    def template(self, request):
        """下载用户导入模板 — GET /api/user/template"""
        import openpyxl
        from io import BytesIO
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "用户导入模板"
        headers = ["用户名", "昵称", "真实姓名", "邮箱", "手机号", "性别", "状态"]
        ws.append(headers)
        ws.append(["zhangsan", "张三", "张三丰", "zhangsan@example.com", "13800138000", "男", "启用"])
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx).font = openpyxl.styles.Font(bold=True)
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=user_template.xlsx"
        wb.save(response)
        return response

    @action(detail=False, methods=["get"])

    def export(self, request):

        """导出用户 — GET /api/user/export"""
        import csv
        from django.http import HttpResponse

        users = self.get_queryset().order_by("-create_time")
        response = HttpResponse(content_type="text/csv; charset=utf-8-sig")
        response["Content-Disposition"] = f"attachment; filename=users_{timezone.now().strftime('%Y%m%d')}.csv"
        response.write(b'\xef\xbb\xbf')
        writer = csv.writer(response)
        writer.writerow(["用户名", "昵称", "邮箱", "手机号", "性别", "状态", "创建时间"])
        for u in users:
            writer.writerow([u.username, u.nickname or "", u.email or "", u.telephone or "", {0: "未知", 1: "男", 2: "女"}.get(u.gender, "未知"), "启用" if u.status == 1 else "禁用", u.create_time])
        return response

    @action(detail=False, methods=["post"], url_path="import")

    def import_(self, request):

        """导入用户 — POST /api/user/import"""
        import openpyxl
        from io import BytesIO

        file = request.FILES.get("file")
        if not file:
            return APIResponse.error(message="请上传文件")
        try:
            wb = openpyxl.load_workbook(BytesIO(file.read()))
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return APIResponse.error(message="文件为空")

            header_map = {
                "用户名": "username", "username": "username",
                "昵称": "nickname", "nickname": "nickname",
                "真实姓名": "real_name", "real_name": "real_name",
                "邮箱": "email", "email": "email",
                "手机号": "telephone", "telephone": "telephone",
                "性别": "gender", "gender": "gender",
                "状态": "status", "status": "status",
            }
            header_row = [str(h).strip() if h else "" for h in rows[0]]
            col_index = {}
            for i, h in enumerate(header_row):
                if h in header_map:
                    col_index[header_map[h]] = i

            if "username" not in col_index:
                return APIResponse.error(message="文件格式错误：未找到用户名列（用户名/username）")

            success = 0
            skipped = 0
            errors = []
            for idx, row in enumerate(rows[1:], start=2):
                cells = list(row)
                username = cells[col_index["username"]] if col_index.get("username") is not None and col_index.get("username") < len(cells) else None
                if not username:
                    skipped += 1
                    errors.append(f"第{idx}行：用户名为空")
                    continue
                username = str(username).strip()

                nickname = str(cells[col_index["nickname"]]).strip() if "nickname" in col_index and col_index["nickname"] < len(cells) and cells[col_index["nickname"]] else ""
                real_name = str(cells[col_index["real_name"]]).strip() if "real_name" in col_index and col_index["real_name"] < len(cells) and cells[col_index["real_name"]] else ""
                email_val = str(cells[col_index["email"]]).strip() if "email" in col_index and col_index["email"] < len(cells) and cells[col_index["email"]] else ""
                telephone = str(cells[col_index["telephone"]]).strip() if "telephone" in col_index and col_index["telephone"] < len(cells) and cells[col_index["telephone"]] else ""
                gender_val = cells[col_index["gender"]] if "gender" in col_index and col_index["gender"] < len(cells) else 0
                status_val = cells[col_index["status"]] if "status" in col_index and col_index["status"] < len(cells) else 1

                if isinstance(gender_val, str):
                    gender_map = {"男": 1, "女": 2, "未知": 0}
                    gender_val = gender_map.get(gender_val, 0)
                elif gender_val is None:
                    gender_val = 0
                if isinstance(status_val, str):
                    status_val = 1 if status_val == "启用" else 0
                elif status_val is None:
                    status_val = 1

                try:
                    user, created = User.objects.update_or_create(
                        username=username,
                        defaults={
                            "nickname": nickname,
                            "real_name": real_name,
                            "email": email_val,
                            "telephone": telephone,
                            "gender": int(gender_val),
                            "status": int(status_val),
                        }
                    )
                    if created:
                        user.set_password("Admin@123")
                        user.save(update_fields=["password", "update_time"])
                    success += 1
                except (IntegrityError, ValueError) as e:
                    skipped += 1
                    errors.append(f"第{idx}行：{str(e)}")

            return APIResponse.success(data={"success": success, "skipped": skipped, "errors": errors}, message=f"导入完成：成功 {success} 条，跳过 {skipped} 条")
        except Exception as e:
            return APIResponse.error(message=f"导入失败：{str(e)}")

    @action(detail=False, methods=["post"])

    def avatar(self, request):

        """

        上传头像 — POST /api/user/avatar

        Request:  multipart/form-data { file: <image> }

        Response: { "code": 200, "data": { "url": "/media/avatars/xxx.jpg" } }

        """

        user = request.user

        if not user or not user.is_authenticated:

            return APIResponse.error(

                message="未登录", code=3000, http_status=401,

            )

        file = request.FILES.get("file")

        if not file:

            return APIResponse.error(

                message="请选择文件", code=2000, http_status=400,

            )

        if file.size > 2 * 1024 * 1024:

            return APIResponse.error(

                message="文件大小不能超过 2MB", code=2000, http_status=400,

            )

        # 校验文件类型

        ext = os.path.splitext(file.name)[1].lower()

        if ext not in (".jpg", ".jpeg", ".png", ".gif", ".webp"):

            return APIResponse.error(

                message="不支持的文件格式，请上传 JPG/PNG/GIF/WebP", code=2000, http_status=400,

            )

        # 保存文件: media/avatars/<uuid><ext>

        avatar_dir = os.path.join(settings.MEDIA_ROOT, "avatars")

        os.makedirs(avatar_dir, exist_ok=True)

        filename = f"{uuid.uuid4().hex}{ext}"

        filepath = os.path.join(avatar_dir, filename)

        with open(filepath, "wb") as f:

            for chunk in file.chunks():

                f.write(chunk)

        # 更新用户头像字段

        url = f"{settings.MEDIA_URL}avatars/{filename}"

        user.avatar = url

        user.save(update_fields=["avatar", "update_time"])

        return APIResponse.success(

            message="上传成功",

            data={"url": url},

        )

    @action(detail=False, methods=["get", "put"])

    def profile(self, request):

        """

        个人资料 — GET/PUT /api/user/profile

        GET  — 返回当前用户信息

        PUT  — 更新个人资料（昵称、邮箱、手机号等）

        """

        user = request.user

        if not user or not user.is_authenticated:

            return APIResponse.error(

                message="未登录", code=3000, http_status=401,

            )

        if request.method == "GET":

            serializer = UserProfileSerializer(user)

            return APIResponse.success(data=serializer.data)

        # PUT — 部分更新

        serializer = UserProfileSerializer(

            user, data=request.data, partial=True,

        )

        if not serializer.is_valid():

            return APIResponse.error(

                message="参数错误", code=2000, http_status=400,

            )

        serializer.save()

        return APIResponse.success(message="更新成功", data=serializer.data)
