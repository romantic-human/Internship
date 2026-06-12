import openpyxl
from django.http import HttpResponse
from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from utils.response import APIResponse
from utils.permissions import HasPermission
from utils.excel import ExcelHandler
from .models import Department
from .serializers import DepartmentSerializer, DepartmentTreeSerializer


class DepartmentViewSet(viewsets.ModelViewSet):
    queryset = Department.objects.all()
    serializer_class = DepartmentSerializer
    permission_key = "dept:list"
    permission_key_map = {
        "create": "dept:add",
        "update": "dept:edit",
        "destroy": "dept:delete",
        "batch": "dept:delete",
        "export": "dept:list",
        "status": "dept:edit",
        "sort": "dept:edit",
        "batch_sort": "dept:edit",
        "template": None,
        "import_departments": "dept:add",
    }

    def get_permissions(self):
        if self.action == "tree":
            return [IsAuthenticated()]
        return [IsAuthenticated(), HasPermission()]

    def perform_create(self, serializer):
        parent_id = serializer.validated_data.pop("parent_id", 0)
        if parent_id:
            serializer.save(parent_id=parent_id)
        else:
            serializer.save(parent=None)

    def perform_update(self, serializer):
        parent_id = serializer.validated_data.pop("parent_id", None)
        if parent_id is not None:
            if parent_id:
                serializer.save(parent_id=parent_id)
            else:
                serializer.save(parent=None)
        else:
            serializer.save()

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        return APIResponse.success(data=serializer.data, message="新增成功")

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(queryset, many=True)
        return APIResponse.success(data=serializer.data)

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return APIResponse.success(data=serializer.data)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop("partial", False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        return APIResponse.success(data=serializer.data, message="更新成功")

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        if Department.objects.filter(parent=instance).exists():
            return APIResponse.error(message="存在子部门，无法删除")
        from apps.user.models import User
        if User.objects.filter(department=instance).exists():
            return APIResponse.error(message="该部门下存在用户，无法删除")
        instance.delete()
        return APIResponse.success(message="删除成功")

    @action(detail=False, methods=["get"], url_path="tree")
    def tree(self, request):
        all_depts = list(Department.objects.all().order_by("sort_order"))
        parent_map = {}
        for d in all_depts:
            d._children = []
            parent_map[d.id] = d
        roots = []
        for d in all_depts:
            if d.parent_id and d.parent_id in parent_map:
                parent_map[d.parent_id]._children.append(d)
            else:
                roots.append(d)
        return APIResponse.success(data=DepartmentTreeSerializer(roots, many=True).data)

    @action(detail=True, methods=["put"], url_path="status")
    def status(self, request, pk=None):
        instance = self.get_object()
        status_val = request.data.get("status")
        if status_val not in (0, 1):
            return APIResponse.error(message="状态值无效")
        instance.status = status_val
        instance.save(update_fields=["status", "update_time"])
        return APIResponse.success(message="状态更新成功")

    @action(detail=True, methods=["put"], url_path="sort")
    def sort(self, request, pk=None):
        instance = self.get_object()
        instance.sort_order = request.data.get("sortOrder", 0)
        instance.save(update_fields=["sort_order", "update_time"])
        return APIResponse.success(message="排序更新成功")

    @action(detail=False, methods=["post"], url_path="batch-sort")
    def batch_sort(self, request):
        data = request.data
        if not isinstance(data, list):
            return APIResponse.error(message="请传入数组")
        instances = []
        for item in data:
            item_id = item.get("id")
            if not item_id:
                return APIResponse.error(message="每项需要 id 字段")
            instances.append(Department(id=item_id, sort_order=item.get("sortOrder", 0)))
        Department.objects.bulk_update(instances, ["sort_order"])
        return APIResponse.success(message="排序更新成功")

    @action(detail=False, methods=["delete"], url_path="batch")
    def batch(self, request):
        ids = request.data.get("ids", [])
        if not ids:
            return APIResponse.error(message="ids 不能为空")
        if Department.objects.filter(parent_id__in=ids).exclude(id__in=ids).exists():
            return APIResponse.error(message="存在子部门不在删除列表中，无法批量删除")
        from apps.user.models import User
        if User.objects.filter(department_id__in=ids).exists():
            return APIResponse.error(message="部分部门下存在用户，无法删除")
        Department.objects.filter(id__in=ids).delete()
        return APIResponse.success(message="批量删除成功")

    @action(detail=False, methods=["get"])
    def export(self, request):
        """\u5bfc\u51fa\u90e8\u95e8"""
        from datetime import datetime
        queryset = Department.objects.all().order_by("sort_order")
        headers = ["\u90e8\u95e8\u540d\u79f0", "\u8d1f\u8d23\u4eba", "\u8054\u7cfb\u7535\u8bdd", "\u90ae\u7bb1", "\u6392\u5e8f", "\u72b6\u6001"]
        rows = [
            [d.dept_name, d.leader or "", d.phone or "", d.email or "", d.sort_order, d.status]
            for d in queryset
        ]
        date_str = datetime.now().strftime("%Y%m%d")
        return ExcelHandler.export_to_response(headers, rows, f"departments_{date_str}.xlsx", "\u90e8\u95e8\u5217\u8868")

    @action(detail=False, methods=["get"], url_path="template")
    def template(self, request):
        """\u4e0b\u8f7d\u90e8\u95e8\u5bfc\u5165\u6a21\u677f"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "\u90e8\u95e8\u5bfc\u5165\u6a21\u677f"
        headers = ["\u90e8\u95e8\u540d\u79f0", "\u7236\u90e8\u95e8\u540d\u79f0", "\u8d1f\u8d23\u4eba", "\u8054\u7cfb\u7535\u8bdd", "\u90ae\u7bb1", "\u6392\u5e8f"]
        ws.append(headers)
        ws.append(["\u6280\u672f\u90e8", "\u603b\u516c\u53f8", "\u5f20\u4e09", "13800000000", "tech@example.com", "10"])
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="department_template.xlsx"'
        wb.save(response)
        return response

    @action(detail=False, methods=["post"], url_path="import-departments")
    def import_departments(self, request):
        """\u5bfc\u5165\u90e8\u95e8"""
        file = request.FILES.get("file")
        if not file:
            return APIResponse.error(message="\u8bf7\u4e0a\u4f20\u6587\u4ef6")
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
        except Exception:
            return APIResponse.error(message="\u6587\u4ef6\u683c\u5f0f\u9519\u8bef")
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        if not rows:
            return APIResponse.error(message="\u6587\u4ef6\u5185\u5bb9\u4e3a\u7a7a")
        success, skipped, errors = 0, 0, []
        all_depts = {d.dept_name: d for d in Department.objects.all()}
        for idx, row in enumerate(rows, start=2):
            dept_name = str(row[0]).strip() if row[0] else ""
            parent_name = str(row[1]).strip() if row[1] else ""
            leader = str(row[2]).strip() if row[2] else ""
            phone = str(row[3]).strip() if row[3] else ""
            email = str(row[4]).strip() if row[4] else ""
            sort_order = int(row[5]) if row[5] is not None else 0
            if not dept_name:
                errors.append(f"\u7b2c{idx}\u884c: \u90e8\u95e8\u540d\u79f0\u4e3a\u7a7a")
                continue
            if dept_name in all_depts:
                skipped += 1
                continue
            parent = all_depts.get(parent_name) if parent_name else None
            Department.objects.create(
                dept_name=dept_name, parent=parent, leader=leader,
                phone=phone, email=email, sort_order=sort_order, status=1,
            )
            success += 1
        return APIResponse.success(data={"success": success, "skipped": skipped, "errors": errors})
