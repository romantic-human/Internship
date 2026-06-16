import openpyxl
from django.http import HttpResponse
from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from utils.response import APIResponse
from utils.permissions import HasPermission
from .models import Menu
from .serializers import MenuSerializer, MenuTreeSerializer


class MenuViewSet(viewsets.ModelViewSet):
    queryset = Menu.objects.all()
    serializer_class = MenuSerializer
    permission_key = "menu:list"
    permission_key_map = {
        "create": "menu:add",
        "update": "menu:edit",
        "destroy": "menu:delete",
        "batch": "menu:delete",
        "export": "menu:list",
        "status": "menu:edit",
        "sort": "menu:edit",
        "batch_sort": "menu:edit",
        "template": None,
        "import_menus": "menu:add",
    }

    def get_permissions(self):
        if self.action in ("tree", "options"):
            return [IsAuthenticated()]
        return [IsAuthenticated(), HasPermission()]

    def get_queryset(self):
        qs = super().get_queryset()
        menu_name = self.request.query_params.get("menu_name")
        if menu_name:
            qs = qs.filter(menu_name__icontains=menu_name)
        return qs

    def perform_create(self, serializer):
        parent_id = serializer.validated_data.pop("parent_id", 0)
        if parent_id:
            serializer.save(parent_id=parent_id)
        else:
            serializer.save(parent=None)

    def perform_update(self, serializer):
        parent_id = serializer.validated_data.pop("parent_id", None)
        if parent_id is not None:
            if parent_id:
                serializer.save(parent_id=parent_id)
            else:
                serializer.save(parent=None)
        else:
            serializer.save()

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        return APIResponse.success(data=serializer.data, message="新增成功")

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(queryset, many=True)
        return APIResponse.success(data=serializer.data)

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return APIResponse.success(data=serializer.data)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop("partial", False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        return APIResponse.success(data=serializer.data, message="更新成功")

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        if Menu.objects.filter(parent=instance).exists():
            return APIResponse.error(message="存在子菜单，无法删除")
        instance.delete()
        return APIResponse.success(message="删除成功")

    @action(detail=False, methods=["get"], url_path="tree")
    def tree(self, request):
        menu_name = request.query_params.get("menu_name", "")
        qs = Menu.objects.all()
        if menu_name:
            qs = qs.filter(menu_name__icontains=menu_name)
        all_menus = list(qs.order_by("sort_order"))
        parent_map = {}
        for m in all_menus:
            m._children = []
            parent_map[m.id] = m
        roots = []
        for m in all_menus:
            if m.parent_id and m.parent_id in parent_map:
                parent_map[m.parent_id]._children.append(m)
            else:
                roots.append(m)
        return APIResponse.success(data=MenuTreeSerializer(roots, many=True).data)

    @action(detail=False, methods=["get"], url_path="options")
    def options(self, request):
        return self.tree(request)

    @action(detail=True, methods=["put"], url_path="status")
    def status(self, request, pk=None):
        instance = self.get_object()
        status_val = request.data.get("status")
        if status_val not in (0, 1):
            return APIResponse.error(message="状态值无效")
        instance.status = status_val
        instance.save(update_fields=["status", "update_time"])
        return APIResponse.success(message="状态更新成功")

    @action(detail=True, methods=["put"], url_path="sort")
    def sort(self, request, pk=None):
        instance = self.get_object()
        sort_order = request.data.get("sortOrder", 0)
        instance.sort_order = sort_order
        instance.save(update_fields=["sort_order", "update_time"])
        return APIResponse.success(message="排序更新成功")

    @action(detail=False, methods=["post"], url_path="batch-sort")
    def batch_sort(self, request):
        data = request.data
        if not isinstance(data, list):
            return APIResponse.error(message="请传入数组")
        instances = []
        for item in data:
            item_id = item.get("id")
            if not item_id:
                return APIResponse.error(message="每项需要 id 字段")
            instances.append(Menu(id=item_id, sort_order=item.get("sortOrder", 0)))
        Menu.objects.bulk_update(instances, ["sort_order"])
        return APIResponse.success(message="排序更新成功")

    @action(detail=False, methods=["delete"], url_path="batch")
    def batch(self, request):
        ids = request.data.get("ids", [])
        if not ids:
            return APIResponse.error(message="ids 不能为空")
        has_children = Menu.objects.filter(parent_id__in=ids).exclude(id__in=ids).exists()
        if has_children:
            return APIResponse.error(message="所选菜单中存在子菜单，无法批量删除")
        Menu.objects.filter(id__in=ids).delete()
        return APIResponse.success(message="批量删除成功")

    @action(detail=False, methods=["get"], url_path="export")
    def export(self, request):
        menus = list(Menu.objects.all().order_by("sort_order"))
        type_map = {0: "目录", 1: "菜单", 2: "按钮"}
        headers = ["ID", "菜单名称", "类型", "路由路径", "组件", "图标", "权限标识", "排序", "状态"]
        rows = [
            [m.id, m.menu_name, type_map.get(m.menu_type, "未知"),
             m.path, m.component, m.icon, m.permission,
             m.sort_order, "启用" if m.status else "禁用"]
            for m in menus
        ]
        from datetime import datetime
        date_str = datetime.now().strftime("%Y%m%d")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "菜单管理"
        ws.append(headers)
        for row in rows:
            ws.append(row)
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="菜单列表_{date_str}.xlsx"'
        wb.save(response)
        return response

    @action(detail=False, methods=["get"], url_path="template")
    def template(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "菜单导入模板"
        headers = ["菜单名称", "类型(0目录/1菜单/2按钮)", "父菜单名称", "路由路径", "组件", "图标", "权限标识", "排序"]
        ws.append(headers)
        ws.append(["系统管理", "0", "", "", "Setting", "", "", "0"])
        ws.append(["用户管理", "1", "系统管理", "/system/user", "system/user/UserList", "User", "user:list", "1"])
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="menu_template.xlsx"'
        wb.save(response)
        return response

    @action(detail=False, methods=["post"], url_path="import-menus")
    def import_menus(self, request):
        file = request.FILES.get("file")
        if not file:
            return APIResponse.error(message="请上传文件")
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
        except Exception:
            return APIResponse.error(message="文件格式错误")
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        if not rows:
            return APIResponse.error(message="文件内容为空")
        success, skipped, errors = 0, 0, []
        all_menus = {m.menu_name: m for m in Menu.objects.all()}
        for idx, row in enumerate(rows, start=2):
            menu_name = str(row[0]).strip() if row[0] else ""
            menu_type = int(row[1]) if row[1] is not None else 0
            parent_name = str(row[2]).strip() if row[2] else ""
            path = str(row[3]).strip() if row[3] else ""
            component = str(row[4]).strip() if row[4] else ""
            icon = str(row[5]).strip() if row[5] else ""
            permission = str(row[6]).strip() if row[6] else ""
            sort_order = int(row[7]) if row[7] is not None else 0
            if not menu_name:
                errors.append(f"第{idx}行: 菜单名称为空")
                continue
            if menu_name in all_menus:
                skipped += 1
                continue
            parent = all_menus.get(parent_name) if parent_name else None
            Menu.objects.create(
                menu_name=menu_name, menu_type=menu_type, parent=parent,
                path=path, component=component, icon=icon,
                permission=permission, sort_order=sort_order,
            )
            success += 1
        return APIResponse.success(data={"success": success, "skipped": skipped, "errors": errors})
