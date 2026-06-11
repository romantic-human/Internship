"""学生中心视图"""
from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from django.db import IntegrityError
from utils.response import APIResponse
from utils.permissions import HasPermission
from .models import StudentInfo, StudentScore
from .serializers import (
    StudentInfoSerializer,
    StudentScoreSerializer,
    StudentScoreCreateSerializer,
)
import csv
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font


class StudentInfoViewSet(viewsets.ModelViewSet):
    """学生信息管理 — /api/student/info/"""

    queryset = StudentInfo.objects.all()
    serializer_class = StudentInfoSerializer
    permission_key = "student:list"
    permission_key_map = {
        "create": "student:add",
        "update": "student:edit",
        "partial_update": "student:edit",
        "destroy": "student:delete",
        "batch": "student:delete",
        "status": "student:edit",
        "export": "student:export",
        "import_students": "student:import",
        "template": None,
    }

    def get_permissions(self):
        return [IsAuthenticated(), HasPermission()]

    def get_queryset(self):
        qs = super().get_queryset()
        name = self.request.query_params.get("name")
        student_no = self.request.query_params.get("student_no")
        class_name = self.request.query_params.get("class_name")
        status_val = self.request.query_params.get("status")
        if name:
            qs = qs.filter(name__icontains=name)
        if student_no:
            qs = qs.filter(student_no__icontains=student_no)
        if class_name:
            qs = qs.filter(class_name__icontains=class_name)
        if status_val is not None and status_val != "":
            qs = qs.filter(status=int(status_val))
        return qs

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = self.get_serializer(queryset, many=True)
        return APIResponse.success(data=serializer.data)

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return APIResponse.success(data=serializer.data)

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return APIResponse.success(data=serializer.data, message="新增成功")

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop("partial", False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return APIResponse.success(data=serializer.data, message="更新成功")

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        if StudentScore.objects.filter(student=instance).exists():
            return APIResponse.error(message="该学生存在关联成绩记录，无法删除")
        instance.delete()
        return APIResponse.success(message="删除成功")

    @action(detail=False, methods=["delete"], url_path="batch")
    def batch(self, request):
        """批量删除学生"""
        ids = request.data.get("ids", [])
        if not ids:
            return APIResponse.error(message="ids 不能为空")
        if StudentScore.objects.filter(student_id__in=ids).exists():
            return APIResponse.error(message="部分学生存在关联成绩记录，无法批量删除")
        StudentInfo.objects.filter(id__in=ids).delete()
        return APIResponse.success(message="批量删除成功")

    @action(detail=True, methods=["put"])
    def status(self, request, pk=None):
        """修改学生状态"""
        instance = self.get_object()
        status_val = request.data.get("status")
        if status_val not in (0, 1, 2):
            return APIResponse.error(message="状态值无效")
        instance.status = status_val
        instance.save(update_fields=["status", "update_time"])
        return APIResponse.success(message="状态更新成功")

    @action(detail=False, methods=["get"], url_path="export")
    def export(self, request):
        """导出学生列表"""
        response = HttpResponse(content_type="text/csv; charset=utf-8-sig")
        response["Content-Disposition"] = 'attachment; filename="students.csv"'
        writer = csv.writer(response)
        writer.writerow(["学号", "姓名", "性别", "班级", "专业", "学院", "手机号", "邮箱", "入学年份", "状态", "创建时间"])
        students = self.get_queryset()
        gender_map = {0: "未知", 1: "男", 2: "女"}
        status_map = {0: "休学", 1: "在读", 2: "毕业"}
        for s in students:
            writer.writerow([
                s.student_no, s.name, gender_map.get(s.gender, "未知"),
                s.class_name, s.major, s.college, s.phone, s.email,
                s.enrollment_year, status_map.get(s.status, "未知"), s.create_time,
            ])
        return response

    @action(detail=False, methods=["get"], url_path="template")
    def template(self, request):
        """下载学生导入模板 — GET /api/student/info/template"""
        wb = Workbook()
        ws = wb.active
        ws.title = "学生导入模板"
        headers = ["学号", "姓名", "性别", "班级", "专业", "学院", "手机号", "邮箱", "入学年份", "状态"]
        ws.append(headers)
        ws.append(["2024001", "张三", "男", "计算机1班", "计算机科学与技术", "信息学院", "13800138000", "zhangsan@example.com", "2024", "在读"])
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=1, column=col_idx).font = Font(bold=True)
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=student_template.xlsx"
        wb.save(response)
        return response

    @action(detail=False, methods=["post"], url_path="import")
    def import_students(self, request):
        """导入学生 — POST /api/student/info/import"""
        from io import BytesIO
        file = request.FILES.get("file")
        if not file:
            return APIResponse.error(message="请上传文件")
        try:
            import openpyxl
            wb = openpyxl.load_workbook(BytesIO(file.read()))
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return APIResponse.error(message="文件为空")

            header_map = {
                "学号": "student_no", "student_no": "student_no",
                "姓名": "name", "name": "name",
                "性别": "gender", "gender": "gender",
                "班级": "class_name", "class_name": "class_name",
                "专业": "major", "major": "major",
                "学院": "college", "college": "college",
                "手机号": "phone", "phone": "phone",
                "邮箱": "email", "email": "email",
                "入学年份": "enrollment_year", "enrollment_year": "enrollment_year",
                "状态": "status", "status": "status",
            }
            header_row = [str(h).strip() if h else "" for h in rows[0]]
            col_index = {}
            for i, h in enumerate(header_row):
                if h in header_map:
                    col_index[header_map[h]] = i

            if "student_no" not in col_index:
                return APIResponse.error(message="文件格式错误：未找到学号列")

            success = 0
            skipped = 0
            errors = []
            existing_nos = set(
                StudentInfo.objects.filter(student_no__in=[
                    str(row[col_index["student_no"]]).strip()
                    for row in rows[1:] if row[col_index["student_no"]]
                ]).values_list("student_no", flat=True)
            ) if col_index.get("student_no") is not None else set()

            for idx, row in enumerate(rows[1:], start=2):
                cells = list(row)
                student_no = cells[col_index["student_no"]] if col_index.get("student_no") is not None and col_index["student_no"] < len(cells) else None
                if not student_no:
                    skipped += 1
                    errors.append(f"第{idx}行：学号为空")
                    continue
                student_no = str(student_no).strip()

                name = str(cells[col_index["name"]]).strip() if "name" in col_index and col_index["name"] < len(cells) and cells[col_index["name"]] else ""
                gender_val = cells[col_index["gender"]] if "gender" in col_index and col_index["gender"] < len(cells) else 0
                class_name = str(cells[col_index["class_name"]]).strip() if "class_name" in col_index and col_index["class_name"] < len(cells) and cells[col_index["class_name"]] else ""
                major = str(cells[col_index["major"]]).strip() if "major" in col_index and col_index["major"] < len(cells) and cells[col_index["major"]] else ""
                college = str(cells[col_index["college"]]).strip() if "college" in col_index and col_index["college"] < len(cells) and cells[col_index["college"]] else ""
                phone = str(cells[col_index["phone"]]).strip() if "phone" in col_index and col_index["phone"] < len(cells) and cells[col_index["phone"]] else ""
                email_val = str(cells[col_index["email"]]).strip() if "email" in col_index and col_index["email"] < len(cells) and cells[col_index["email"]] else ""
                enrollment_year = str(cells[col_index["enrollment_year"]]).strip() if "enrollment_year" in col_index and col_index["enrollment_year"] < len(cells) and cells[col_index["enrollment_year"]] else ""
                status_val = cells[col_index["status"]] if "status" in col_index and col_index["status"] < len(cells) else 1

                if isinstance(gender_val, str):
                    gender_map = {"男": 1, "女": 2, "未知": 0}
                    gender_val = gender_map.get(gender_val, 0)
                elif gender_val is None:
                    gender_val = 0
                else:
                    try:
                        gender_val = int(gender_val)
                    except (ValueError, TypeError):
                        gender_val = 0

                if isinstance(status_val, str):
                    status_map = {"在读": 1, "休学": 0, "毕业": 2}
                    status_val = status_map.get(status_val, 1)
                elif status_val is None:
                    status_val = 1
                else:
                    try:
                        status_val = int(status_val)
                    except (ValueError, TypeError):
                        status_val = 1

                if student_no in existing_nos:
                    skipped += 1
                    errors.append(f"第{idx}行：学号 {student_no} 已存在")
                    continue

                try:
                    StudentInfo.objects.create(
                        student_no=student_no, name=name,
                        gender=gender_val, class_name=class_name,
                        major=major, college=college,
                        phone=phone, email=email_val,
                        enrollment_year=enrollment_year,
                        status=status_val,
                    )
                    success += 1
                except IntegrityError as e:
                    skipped += 1
                    errors.append(f"第{idx}行：{str(e)}")

            message = f"导入完成：成功 {success} 条"
            if skipped:
                message += f"，跳过 {skipped} 条"
            return APIResponse.success(data={"success": success, "skipped": skipped, "errors": errors[:100]}, message=message)
        except Exception as e:
            return APIResponse.error(message=f"导入失败：{str(e)}")


class StudentScoreViewSet(viewsets.ModelViewSet):
    """学生成绩管理 — /api/student/score/"""

    queryset = StudentScore.objects.select_related("student").all()
    serializer_class = StudentScoreSerializer
    permission_key = "score:list"
    permission_key_map = {
        "create": "score:add",
        "update": "score:edit",
        "partial_update": "score:edit",
        "destroy": "score:delete",
        "batch": "score:delete",
        "export": "score:export",
        "import_scores": "score:import",
        "template": None,
    }

    def get_permissions(self):
        return [IsAuthenticated(), HasPermission()]

    def get_serializer_class(self):
        if self.action in ("create", "update", "partial_update"):
            return StudentScoreCreateSerializer
        return StudentScoreSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        student_id = self.request.query_params.get("student_id")
        course_name = self.request.query_params.get("course_name")
        semester = self.request.query_params.get("semester")
        if student_id:
            qs = qs.filter(student_id=int(student_id))
        if course_name:
            qs = qs.filter(course_name__icontains=course_name)
        if semester:
            qs = qs.filter(semester=semester)
        return qs

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = self.get_serializer(queryset, many=True)
        return APIResponse.success(data=serializer.data)

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return APIResponse.success(data=serializer.data)

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return APIResponse.success(data=serializer.data, message="新增成功")

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop("partial", False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return APIResponse.success(data=serializer.data, message="更新成功")

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.delete()
        return APIResponse.success(message="删除成功")

    @action(detail=False, methods=["delete"], url_path="batch")
    def batch(self, request):
        """批量删除成绩"""
        ids = request.data.get("ids", [])
        if not ids:
            return APIResponse.error(message="ids 不能为空")
        StudentScore.objects.filter(id__in=ids).delete()
        return APIResponse.success(message="批量删除成功")

    @action(detail=False, methods=["get"], url_path="export")
    def export(self, request):
        """导出成绩列表"""
        response = HttpResponse(content_type="text/csv; charset=utf-8-sig")
        response["Content-Disposition"] = 'attachment; filename="scores.csv"'
        writer = csv.writer(response)
        writer.writerow(["学号", "姓名", "课程名称", "成绩", "学分", "学期", "创建时间"])
        scores = self.get_queryset()
        for s in scores:
            writer.writerow([
                s.student.student_no, s.student.name, s.course_name,
                s.score, s.credit, s.semester, s.create_time,
            ])
        return response

    @action(detail=False, methods=["get"], url_path="template")
    def template(self, request):
        """下载成绩导入模板 — GET /api/student/score/template"""
        wb = Workbook()
        ws = wb.active
        ws.title = "成绩导入模板"
        headers = ["学号", "课程名称", "成绩", "学分", "学期"]
        ws.append(headers)
        ws.append(["2024001", "数据结构", 90, 4, "2024-2025-1"])
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=1, column=col_idx).font = Font(bold=True)
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=score_template.xlsx"
        wb.save(response)
        return response

    @action(detail=False, methods=["post"], url_path="import")
    def import_scores(self, request):
        """导入成绩 — POST /api/student/score/import"""
        import openpyxl
        from io import BytesIO
        file = request.FILES.get("file")
        if not file:
            return APIResponse.error(message="请上传文件")
        try:
            wb = openpyxl.load_workbook(BytesIO(file.read()))
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return APIResponse.error(message="文件为空")

            header_map = {
                "学号": "student_no", "student_no": "student_no",
                "课程名称": "course_name", "course_name": "course_name",
                "成绩": "score", "score": "score",
                "学分": "credit", "credit": "credit",
                "学期": "semester", "semester": "semester",
            }
            header_row = [str(h).strip() if h else "" for h in rows[0]]
            col_index = {}
            for i, h in enumerate(header_row):
                if h in header_map:
                    col_index[header_map[h]] = i

            if "student_no" not in col_index:
                return APIResponse.error(message="文件格式错误：未找到学号列")

            success = 0
            skipped = 0
            errors = []
            students_map = {}
            if col_index.get("student_no") is not None:
                student_nos_for_lookup = [str(row[col_index["student_no"]]).strip() for row in rows[1:] if row[col_index["student_no"]]]
                if student_nos_for_lookup:
                    students_map = {
                        s.student_no: s
                        for s in StudentInfo.objects.filter(student_no__in=student_nos_for_lookup).only("id", "student_no")
                    }

            for idx, row in enumerate(rows[1:], start=2):
                cells = list(row)
                student_no = cells[col_index["student_no"]] if col_index.get("student_no") is not None and col_index["student_no"] < len(cells) else None
                if not student_no:
                    skipped += 1
                    errors.append(f"第{idx}行：学号为空")
                    continue
                student_no = str(student_no).strip()

                course_name = str(cells[col_index["course_name"]]).strip() if "course_name" in col_index and col_index["course_name"] < len(cells) and cells[col_index["course_name"]] else ""
                if not course_name:
                    skipped += 1
                    errors.append(f"第{idx}行：课程名称为空")
                    continue

                score_val = cells[col_index["score"]] if "score" in col_index and col_index["score"] < len(cells) else 0
                credit_val = cells[col_index["credit"]] if "credit" in col_index and col_index["credit"] < len(cells) else 0
                semester = str(cells[col_index["semester"]]).strip() if "semester" in col_index and col_index["semester"] < len(cells) and cells[col_index["semester"]] else ""

                try:
                    score_val = float(score_val) if score_val else 0
                except (ValueError, TypeError):
                    score_val = 0
                try:
                    credit_val = float(credit_val) if credit_val else 0
                except (ValueError, TypeError):
                    credit_val = 0

                student = students_map.get(student_no)
                if not student:
                    skipped += 1
                    errors.append(f"第{idx}行：学号 {student_no} 不存在")
                    continue

                try:
                    StudentScore.objects.create(
                        student=student, course_name=course_name,
                        score=score_val, credit=credit_val,
                        semester=semester,
                    )
                    success += 1
                except IntegrityError as e:
                    skipped += 1
                    errors.append(f"第{idx}行：{str(e)}")

            message = f"导入完成：成功 {success} 条"
            if skipped:
                message += f"，跳过 {skipped} 条"
            return APIResponse.success(data={"success": success, "skipped": skipped, "errors": errors[:100]}, message=message)
        except Exception as e:
            return APIResponse.error(message=f"导入失败：{str(e)}")
