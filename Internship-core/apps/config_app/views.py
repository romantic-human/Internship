import os
import uuid

from django.conf import settings
from django.db import IntegrityError, transaction
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import viewsets
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from utils.response import APIResponse
from utils.permissions import HasPermission
from .models import SystemConfig
from .serializers import SystemConfigSerializer

PANEL_DEFAULTS = {
    "system.name": "企业智能分析平台",
    "system.logo": "",
    "log.enabled": "1",
    "log.retention_days": "90",
    "log.alert_enabled": "1",
    "security.level": "高",
    "security.two_factor": "1",
    "security.password_policy": "至少8位，包含数字与大小写字母",
}


class SystemConfigViewSet(viewsets.ModelViewSet):
    queryset = SystemConfig.objects.all()
    serializer_class = SystemConfigSerializer
    permission_key = "config:list"
    permission_key_map = {
        "create": "config:add",
        "update": "config:edit",
        "partial_update": "config:edit",
        "destroy": "config:delete",
        "batch": "config:delete",
        "export": "config:export",
        "status": "config:edit",
        "sort": "config:edit",
        "batch_sort": "config:edit",
        "panel_save": "config:edit",
        "import_configs": "config:add",
        "template": None,
    }

    def get_permissions(self):
        if self.action in ("by_key", "panel_get"):
            return [IsAuthenticated()]
        return [IsAuthenticated(), HasPermission()]

    def get_queryset(self):
        qs = super().get_queryset()
        config_name = self.request.query_params.get("config_name")
        config_key = self.request.query_params.get("config_key")
        if config_name:
            qs = qs.filter(config_name__icontains=config_name)
        if config_key:
            qs = qs.filter(config_key__icontains=config_key)
        return qs

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return APIResponse.success(data=serializer.data, message="新增成功")

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = self.get_serializer(queryset, many=True)
        return APIResponse.success(data=serializer.data)

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return APIResponse.success(data=serializer.data)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop("partial", False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return APIResponse.success(data=serializer.data, message="更新成功")

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.delete()
        return APIResponse.success(message="删除成功")

    @action(detail=False, methods=["delete"], url_path="batch")
    def batch(self, request):
        ids = request.data.get("ids", [])
        if not ids:
            return APIResponse.error(message="ids 不能为空")
        SystemConfig.objects.filter(id__in=ids).delete()
        return APIResponse.success(message="批量删除成功")

    @action(detail=False, methods=["get"], url_path="export")
    def export(self, request):
        configs = self.get_queryset()
        response = HttpResponse(content_type="text/csv; charset=utf-8-sig")
        response["Content-Disposition"] = 'attachment; filename="configs.csv"'
        import csv
        writer = csv.writer(response)
        writer.writerow(["配置键", "配置名称", "配置值", "配置类型", "排序", "状态", "创建时间"])
        type_map = {0: "字符串", 1: "数字", 2: "布尔", 3: "JSON"}
        status_map = {0: "禁用", 1: "启用"}
        for c in configs:
            writer.writerow([
                c.config_key, c.config_name, c.config_value,
                type_map.get(c.config_type, c.config_type),
                c.sort_order, status_map.get(c.status, c.status),
                c.create_time,
            ])
        return response

    @action(detail=False, methods=["get"], url_path="by-key")
    def by_key(self, request):
        key = request.query_params.get("key", "")
        if not key:
            return APIResponse.error(message="缺少参数 key")
        config = SystemConfig.objects.filter(config_key=key).first()
        if config:
            return APIResponse.success(data=config.config_value)
        return APIResponse.not_found()

    @action(detail=True, methods=["put"], url_path="sort")
    def sort(self, request, pk=None):
        instance = self.get_object()
        instance.sort_order = request.data.get("sortOrder", 0)
        instance.save(update_fields=["sort_order", "update_time"])
        return APIResponse.success(message="排序更新成功")

    @action(detail=True, methods=["put"])
    def status(self, request, pk=None):
        instance = self.get_object()
        status_val = request.data.get("status")
        if status_val not in (0, 1):
            return APIResponse.error(message="状态值无效")
        instance.status = status_val
        instance.save(update_fields=["status", "update_time"])
        return APIResponse.success(message="状态更新成功")

    @action(detail=False, methods=["post"], url_path="batch-sort")
    def batch_sort(self, request):
        data = request.data
        if not isinstance(data, list):
            return APIResponse.error(message="请传入数组")
        instances = []
        for item in data:
            item_id = item.get("id")
            if not item_id:
                return APIResponse.error(message="每项需要 id 字段")
            instances.append(SystemConfig(id=item_id, sort_order=item.get("sortOrder", 0)))
        now = timezone.now()
        for inst in instances:
            inst.update_time = now
        SystemConfig.objects.bulk_update(instances, ["sort_order", "update_time"])
        return APIResponse.success(message="排序更新成功")

    @action(detail=False, methods=["get"], url_path="panel")
    def panel_get(self, request):
        configs = SystemConfig.objects.filter(config_key__in=PANEL_DEFAULTS.keys())
        data = {c.config_key: c.config_value for c in configs}
        for key, default in PANEL_DEFAULTS.items():
            if key not in data:
                data[key] = default
        return APIResponse.success(data=data)

    @action(detail=False, methods=["post"], url_path="panel-save")
    def panel_save(self, request):
        payload = request.data
        with transaction.atomic():
            for key, value in payload.items():
                if key not in PANEL_DEFAULTS:
                    continue
                obj, created = SystemConfig.objects.get_or_create(
                    config_key=key,
                    defaults={"config_name": key, "config_value": str(value), "config_type": 0, "status": 1},
                )
                if not created:
                    obj.config_value = str(value)
                    obj.save(update_fields=["config_value", "update_time"])
        return APIResponse.success(message="配置保存成功")

    @action(detail=False, methods=["get"], url_path="template")
    def template(self, request):
        """下载配置导入模板 — GET /api/config/template"""
        from openpyxl import Workbook
        from openpyxl.styles import Font
        wb = Workbook()
        ws = wb.active
        ws.title = "配置导入模板"
        headers = ["配置键", "配置名称", "配置值", "配置类型", "排序", "状态"]
        ws.append(headers)
        ws.append(["site.title", "网站标题", "我的站点", "0", "1", "1"])
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=1, column=col_idx).font = Font(bold=True)
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=config_template.xlsx"
        wb.save(response)
        return response

    @action(detail=False, methods=["post"], url_path="import")
    def import_configs(self, request):
        """导入配置 — POST /api/config/import"""
        from io import BytesIO
        import openpyxl
        file = request.FILES.get("file")
        if not file:
            return APIResponse.error(message="请上传文件")
        try:
            wb = openpyxl.load_workbook(BytesIO(file.read()))
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return APIResponse.error(message="文件为空")

            header_map = {
                "配置键": "config_key", "config_key": "config_key",
                "配置名称": "config_name", "config_name": "config_name",
                "配置值": "config_value", "config_value": "config_value",
                "配置类型": "config_type", "config_type": "config_type",
                "排序": "sort_order", "sort_order": "sort_order",
                "状态": "status", "status": "status",
            }
            header_row = [str(h).strip() if h else "" for h in rows[0]]
            col_index = {}
            for i, h in enumerate(header_row):
                if h in header_map:
                    col_index[header_map[h]] = i

            if "config_key" not in col_index:
                return APIResponse.error(message="文件格式错误：未找到配置键列")

            success = 0
            skipped = 0
            errors = []
            for idx, row in enumerate(rows[1:], start=2):
                cells = list(row)
                config_key = cells[col_index["config_key"]] if col_index.get("config_key") is not None and col_index["config_key"] < len(cells) else None
                if not config_key:
                    skipped += 1
                    errors.append(f"第{idx}行：配置键为空")
                    continue
                config_key = str(config_key).strip()

                config_name = str(cells[col_index["config_name"]]).strip() if "config_name" in col_index and col_index["config_name"] < len(cells) and cells[col_index["config_name"]] else ""
                config_value = str(cells[col_index["config_value"]]).strip() if "config_value" in col_index and col_index["config_value"] < len(cells) and cells[col_index["config_value"]] else ""
                config_type = cells[col_index["config_type"]] if "config_type" in col_index and col_index["config_type"] < len(cells) else 0
                sort_order = cells[col_index["sort_order"]] if "sort_order" in col_index and col_index["sort_order"] < len(cells) else 0
                status_val = cells[col_index["status"]] if "status" in col_index and col_index["status"] < len(cells) else 1

                try:
                    config_type = int(config_type) if config_type else 0
                except (ValueError, TypeError):
                    config_type = 0
                try:
                    sort_order = int(sort_order) if sort_order else 0
                except (ValueError, TypeError):
                    sort_order = 0
                try:
                    status_val = int(status_val) if status_val else 1
                except (ValueError, TypeError):
                    status_val = 1

                if SystemConfig.objects.filter(config_key=config_key).exists():
                    skipped += 1
                    errors.append(f"第{idx}行：配置键 {config_key} 已存在")
                    continue

                try:
                    SystemConfig.objects.create(
                        config_key=config_key, config_name=config_name,
                        config_value=config_value, config_type=config_type,
                        sort_order=sort_order, status=status_val,
                    )
                    success += 1
                except IntegrityError as e:
                    skipped += 1
                    errors.append(f"第{idx}行：{str(e)}")

            message = f"导入完成：成功 {success} 条"
            if skipped:
                message += f"，跳过 {skipped} 条"
            return APIResponse.success(data={"success": success, "skipped": skipped, "errors": errors[:100]}, message=message)
        except Exception as e:
            return APIResponse.error(message=f"导入失败：{str(e)}")


@api_view(["POST"])
@permission_classes([IsAuthenticated, HasPermission])
def upload_image(request):
    upload_image.permission_key = "config:edit"
    file = request.FILES.get("file")
    if not file:
        return APIResponse.error(message="请选择文件", code=2000, http_status=400)
    if file.size > 2 * 1024 * 1024:
        return APIResponse.error(message="文件大小不能超过 2MB", code=2000, http_status=400)
    ext = os.path.splitext(file.name)[1].lower()
    if ext not in (".jpg", ".jpeg", ".png", ".gif", ".svg", ".webp"):
        return APIResponse.error(message="不支持的格式", code=2000, http_status=400)
    upload_dir = os.path.join(settings.MEDIA_ROOT, "uploads")
    os.makedirs(upload_dir, exist_ok=True)
    filename = f"{uuid.uuid4().hex}{ext}"
    filepath = os.path.join(upload_dir, filename)
    with open(filepath, "wb") as f:
        for chunk in file.chunks():
            f.write(chunk)
    url = f"{settings.MEDIA_URL}uploads/{filename}"
    return APIResponse.success(message="上传成功", data={"url": url})
