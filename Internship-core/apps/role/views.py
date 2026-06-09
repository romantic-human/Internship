"""角色模块视图 — 参考《组织架构模块设计方案.md》第 5.3 节"""
from rest_framework import viewsets
from rest_framework.decorators import action
from django.db import transaction
from utils.response import APIResponse
from .models import Role, RoleMenuRelation
from .serializers import (
    RoleSerializer,
    AssignMenuSerializer,
    AssignUserSerializer,
)
from apps.menu.models import Menu
from apps.user.models import User, UserRoleRelation
import csv
import openpyxl
from io import BytesIO
from django.http import HttpResponse
from rest_framework.permissions import IsAuthenticated
from utils.permissions import HasPermission


class RoleViewSet(viewsets.ModelViewSet):
    queryset = Role.objects.all()
    serializer_class = RoleSerializer
    permission_key = "role:list"
    permission_key_map = {
        "menus": "role:assign",
        "users": "role:assign",
        "batch": "role:delete",
    }

    def get_permissions(self):
        return [IsAuthenticated(), HasPermission()]

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return APIResponse.success(data=serializer.data, message="新增成功")

    def get_queryset(self):
        qs = super().get_queryset()
        role_name = self.request.query_params.get("role_name")
        status_val = self.request.query_params.get("status")
        if role_name:
            qs = qs.filter(role_name__icontains=role_name)
        if status_val is not None and status_val != "":
            qs = qs.filter(status=int(status_val))
        return qs

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = self.get_serializer(queryset, many=True)
        return APIResponse.success(data=serializer.data)

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return APIResponse.success(data=serializer.data)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop("partial", False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return APIResponse.success(data=serializer.data, message="更新成功")

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        # 检查是否有关联用户
        if UserRoleRelation.objects.filter(role=instance).exists():
            return APIResponse.error(message="该角色下存在关联用户，无法删除")
        instance.delete()
        return APIResponse.success(message="删除成功")

    @action(detail=False, methods=["delete"], url_path="batch")
    def batch(self, request):
        """批量删除 — DELETE /api/role/batch"""
        ids = request.data.get("ids", [])
        if not ids:
            return APIResponse.error(message="ids 不能为空")
        # 检查是否有关联用户
        if UserRoleRelation.objects.filter(role_id__in=ids).exists():
            return APIResponse.error(message="部分角色存在关联用户，无法批量删除")
        Role.objects.filter(id__in=ids).delete()
        return APIResponse.success(message="批量删除成功")

    @action(detail=False, methods=["get"])
    def all(self, request):
        """获取全部角色（下拉框用）"""
        roles = Role.objects.filter(status=1).order_by("role_sort")
        serializer = RoleSerializer(roles, many=True)
        return APIResponse.success(data=serializer.data)

    @action(detail=True, methods=["put"])
    def status(self, request, pk=None):
        instance = self.get_object()
        status_val = request.data.get("status")
        if status_val not in (0, 1):
            return APIResponse.error(message="状态值无效")
        instance.status = status_val
        instance.save()
        return APIResponse.success(message="状态更新成功")

    @action(detail=True, methods=["put"], url_path="sort")
    def sort(self, request, pk=None):
        instance = self.get_object()
        instance.role_sort = request.data.get("sortOrder", 0)
        instance.save()
        return APIResponse.success(message="排序更新成功")

    @action(detail=False, methods=["post"], url_path="batch-sort")
    def batch_sort(self, request):
        data = request.data
        if not isinstance(data, list):
            return APIResponse.error(message="请传入列表")
        instances = []
        for item in data:
            item_id = item.get("id")
            if not item_id:
                return APIResponse.error(message="每项需要 id 字段")
            instances.append(Role(id=item_id, role_sort=item.get("sortOrder", 0)))
        Role.objects.bulk_update(instances, ["role_sort"])
        return APIResponse.success(message="批量排序成功")

    @action(detail=False, methods=["get"], url_path="template")
    def template(self, request):
        """下载角色导入模板 — GET /api/role/template"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "角色导入模板"
        headers = ["角色名称", "角色标识", "排序", "状态", "备注"]
        ws.append(headers)
        ws.append(["管理员", "admin", 0, "启用", "系统管理员角色"])
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx).font = openpyxl.styles.Font(bold=True)
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=role_template.xlsx"
        wb.save(response)
        return response

    @action(detail=False, methods=["get"], url_path="export")
    def export(self, request):
        response = HttpResponse(content_type="text/csv; charset=utf-8-sig")
        response["Content-Disposition"] = 'attachment; filename="roles.csv"'
        writer = csv.writer(response)
        writer.writerow(["ID", "角色名称", "角色标识", "排序", "状态", "创建时间"])
        roles = Role.objects.all().order_by("role_sort")
        for r in roles:
            writer.writerow([r.id, r.role_name, r.role_key, r.role_sort, "启用" if r.status else "禁用", r.create_time])
        return response

    @action(detail=False, methods=["post"], url_path="import")
    def import_roles(self, request):
        """导入角色 — POST /api/role/import"""
        file = request.FILES.get("file")
        if not file:
            return APIResponse.error(message="请上传文件")
        try:
            wb = openpyxl.load_workbook(BytesIO(file.read()))
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return APIResponse.error(message="文件为空")

            # 表头映射（支持中文表头和英文字段名）
            header_map = {
                "角色名称": "role_name", "role_name": "role_name",
                "角色标识": "role_key", "role_key": "role_key",
                "排序": "role_sort", "role_sort": "role_sort",
                "状态": "status", "status": "status",
                "备注": "remark", "remark": "remark",
            }
            header_row = [str(h).strip() if h else "" for h in rows[0]]
            col_index = {}
            for i, h in enumerate(header_row):
                if h in header_map:
                    col_index[header_map[h]] = i

            # 如果没有识别到表头，按默认顺序
            if not col_index:
                col_index = {"role_name": 0, "role_key": 1, "role_sort": 2, "status": 3, "remark": 4}
                data_rows = rows[1:]
            else:
                data_rows = rows[1:]

            # 校验必须包含角色名称和角色标识列
            if "role_name" not in col_index or "role_key" not in col_index:
                return APIResponse.error(message="文件格式错误：未找到角色名称或角色标识列，请导入角色数据文件")

            count = 0
            for row in data_rows:
                cells = list(row)
                role_name = cells[col_index.get("role_name", 0)] if col_index.get("role_name") is not None and col_index.get("role_name") < len(cells) else None
                role_key = cells[col_index.get("role_key", 1)] if col_index.get("role_key") is not None and col_index.get("role_key") < len(cells) else None
                if not role_name or not role_key:
                    continue
                role_sort = cells[col_index["role_sort"]] if "role_sort" in col_index and col_index["role_sort"] < len(cells) else 0
                status_val = cells[col_index["status"]] if "status" in col_index and col_index["status"] < len(cells) else 1
                remark = cells[col_index["remark"]] if "remark" in col_index and col_index["remark"] < len(cells) else ""

                # 类型转换
                try:
                    role_sort = int(role_sort) if role_sort else 0
                except (ValueError, TypeError):
                    role_sort = 0
                if isinstance(status_val, str):
                    status_val = 1 if status_val == "启用" else 0
                elif status_val is None:
                    status_val = 1

                Role.objects.update_or_create(
                    role_key=str(role_key).strip(),
                    defaults={
                        "role_name": str(role_name).strip(),
                        "role_sort": role_sort,
                        "status": int(status_val),
                        "remark": str(remark or "").strip(),
                    }
                )
                count += 1
            return APIResponse.success(message=f"导入成功，共 {count} 条")
        except Exception as e:
            return APIResponse.error(message=f"导入失败: {str(e)}")

    @action(detail=True, methods=["get", "put"], url_path="menus")
    def menus(self, request, pk=None):
        """获取/分配角色菜单权限"""
        instance = self.get_object()

        if request.method == "GET":
            menu_ids = RoleMenuRelation.objects.filter(
                role=instance
            ).values_list("menu_id", flat=True)
            return APIResponse.success(data=list(menu_ids))

        # PUT — 分配菜单
        serializer = AssignMenuSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        menu_ids = serializer.validated_data["menu_ids"]

        with transaction.atomic():
            RoleMenuRelation.objects.filter(role=instance).delete()
            if menu_ids:
                menus = Menu.objects.filter(id__in=menu_ids)
                relations = [
                    RoleMenuRelation(role=instance, menu=menu)
                    for menu in menus
                ]
                RoleMenuRelation.objects.bulk_create(relations)

        return APIResponse.success(message="菜单权限分配成功")

    @action(detail=True, methods=["get", "put"], url_path="users")
    def users(self, request, pk=None):
        """获取/分配角色下的用户"""
        instance = self.get_object()

        if request.method == "GET":
            user_ids = UserRoleRelation.objects.filter(
                role=instance
            ).values_list("user_id", flat=True)
            return APIResponse.success(data=list(user_ids))

        # PUT — 分配用户
        serializer = AssignUserSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user_ids = serializer.validated_data["user_ids"]

        with transaction.atomic():
            UserRoleRelation.objects.filter(role=instance).delete()
            if user_ids:
                users = User.objects.filter(id__in=user_ids)
                relations = [
                    UserRoleRelation(role=instance, user=user)
                    for user in users
                ]
                UserRoleRelation.objects.bulk_create(relations)

        return APIResponse.success(message="用户分配成功")